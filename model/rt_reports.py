"""Refresh the tracked report docs to the new 432-image retrain:
  - REPORT.md          (paper-style narrative + results tables + figure index)
  - LABEL_AUDIT.md     (annotation-vs-folder label audit on the current set)
  - misclassified_6sheets.xlsx  (per-model misclassified TEST images workbook)
"""
from __future__ import annotations
import json
from collections import Counter
from pathlib import Path
import numpy as np

import du

REPO = du.ROOT.parent
PROJ = REPO.parent
MET = json.loads((du.ROOT / "outputs_final" / "metrics_full.json").read_text())
D, MODELS = MET["dataset"], MET["models"]
LABELS = json.loads((PROJ / "data" / "label_map.json").read_text())
ORDER = ["Ensemble (best)", "ResNet50", "ConvNeXt-Tiny", "EfficientNet-B0", "ResNet18",
         "CNN", "SVM (RBF)", "Logistic Regression", "Naive Bayes"]
SHORT = {"Ensemble (best)": "ensemble", "ResNet50": "resnet50", "ConvNeXt-Tiny": "convnext",
         "EfficientNet-B0": "efficientnet", "ResNet18": "resnet18", "CNN": "cnn",
         "SVM (RBF)": "svm", "Logistic Regression": "logreg", "Naive Bayes": "nb"}


def bal(m, sp):
    return MODELS[m][sp]["balanced"]


# ───────────────────────── REPORT.md ─────────────────────────
def report_md():
    e = bal("Ensemble (best)", "test")
    hr = MODELS["Ensemble (best)"]["test"]["high_recall"]
    mr = MODELS["Ensemble (best)"]["test"]["max_recall"]
    L = []
    A = L.append
    A("# Hanging-Passenger (Safe / Unsafe) Classifier — Technical Report\n")
    A("Detecting passengers hanging on the doors of **buses / legunas** — a Dhaka "
      "road-safety violation — as a binary image-classification task. This report "
      "reflects the **432-image** retrain.\n")
    A("## 1. Dataset\n")
    A(f"- **{D['total_labeled_images']} annotated images**, labels derived from the "
      "image **annotations** (a box labelled `unsafe` ⇒ unsafe; else `safe`; `license` ignored) — "
      "not the bucket folder names.\n"
      f"- Class balance: **292 safe / 140 unsafe** (≈2.1 : 1).\n"
      f"- Source: Cloudflare R2 bucket `machine-learning` (raw images + annotations).\n")
    sc = D["split_counts"]
    A("## 2. Method — split & augmentation (no leakage)\n")
    A(f"- **4-way stratified (vehicle × class) 70 / 15 / 15** split so val & test each "
      "carry every category (bus-safe, bus-unsafe, legua-safe, legua-unsafe).\n"
      f"- **Train** = {D['train_total']} images = {D['train_originals']} originals + "
      f"**{D['train_augmented_copies']} offline A–Z augmentations**, class-balanced "
      f"({sc['train']['0']} safe / {sc['train']['1']} unsafe).\n"
      f"- **Val** = {D['val_total']}, **Test** = {D['test_total']} — real originals only "
      "(augmentation applied to TRAIN only → no data leakage).\n"
      "- Deep nets: ImageNet-pretrained, two-phase fine-tune, online aug + WeightedRandomSampler, "
      "threshold tuned on val, hflip TTA. Classical: HOG → StandardScaler → PCA → classifier. "
      "GPU = RTX 5090.\n")
    A("## 3. Results — held-out TEST (balanced operating point)\n")
    hdr = ["Model", "Acc", "Bal-Acc", "Recall (unsafe)", "Precision", "F1", "ROC-AUC", "PR-AUC", "MCC"]
    A("| " + " | ".join(hdr) + " |")
    A("| " + " | ".join(["---"] * len(hdr)) + " |")
    for m in ORDER:
        b = bal(m, "test")
        A(f"| **{m}** | {b['accuracy']:.3f} | {b['balanced_accuracy']:.3f} | "
          f"{b['recall_unsafe']:.3f} | {b['precision_unsafe']:.3f} | {b['f1_unsafe']:.3f} | "
          f"{b['roc_auc']:.3f} | {b['pr_auc']:.3f} | {b['mcc']:.3f} |")
    A("")
    A("## 4. Train / Val / Test accuracy (generalization)\n")
    A("| Model | Train | Val | Test | Train→Test gap |")
    A("| --- | --- | --- | --- | --- |")
    for m in ORDER:
        tr, va, te = bal(m, "train")["accuracy"], bal(m, "val")["accuracy"], bal(m, "test")["accuracy"]
        A(f"| {m} | {tr:.3f} | {va:.3f} | {te:.3f} | {100*(tr-te):.1f} pts |")
    A("")
    A("## 5. Deployed operating modes (Ensemble = ResNet50 + ConvNeXt-Tiny)\n")
    A("| Mode | Test acc | Unsafe recall | Use when |")
    A("| --- | --- | --- | --- |")
    A(f"| Balanced | {e['accuracy']*100:.1f}% | {e['recall_unsafe']*100:.1f}% | max overall accuracy |")
    A(f"| High-recall (default) | {hr['accuracy']*100:.1f}% | {hr['recall_unsafe']*100:.1f}% | safety-leaning |")
    A(f"| Max-recall (zero-miss) | {mr['accuracy']*100:.1f}% | {mr['recall_unsafe']*100:.1f}% | never miss an unsafe |")
    A("")
    A("## 6. Headline\n")
    A(f"**Ensemble — TEST accuracy {e['accuracy']*100:.1f}%, unsafe-recall "
      f"{e['recall_unsafe']*100:.1f}%, ROC-AUC {e['roc_auc']:.3f}.** The added unsafe data "
      f"raised AUC past the old 0.929 ceiling and lifted the zero-miss (100%-recall) "
      f"operating point to ~{mr['accuracy']*100:.0f}% accuracy (was ~26%).\n")
    A("## 7. Figures & tables\n")
    A("See `paper_assets/figures/` (fig01–fig12) and `paper_assets/tables/`. "
      "Full per-metric data: `model/outputs_final/metrics_full.json`.\n")
    (REPO / "REPORT.md").write_text("\n".join(L))
    print("wrote REPORT.md")


# ───────────────────────── LABEL_AUDIT.md ─────────────────────────
def label_audit_md():
    mism = []
    for r in LABELS:
        folder_lbl = "unsafe" if r["folder_class"] == "positive" else "safe"
        if r["label"] != folder_lbl:
            mism.append(r)
    by = Counter(r["label"] for r in LABELS)
    L = ["# Label Audit — annotation-derived vs bucket folder\n",
         f"- Total labeled images: **{len(LABELS)}** (safe {by['safe']} / unsafe {by['unsafe']}).",
         "- Labels are ground truth from **annotations** (a box labelled `unsafe` ⇒ unsafe; "
         "else `safe`; `license` ignored), which can differ from the R2 bucket folder name.",
         f"- **{len(mism)} image(s)** have an annotation label that disagrees with their folder.\n"]
    if mism:
        L.append("| Image | Vehicle | Bucket folder | Annotation label |")
        L.append("| --- | --- | --- | --- |")
        for r in sorted(mism, key=lambda r: (r["vehicle"], r["stem"])):
            L.append(f"| {r['stem']} | {r['vehicle']} | {r['folder_class']} | **{r['label']}** |")
    else:
        L.append("_No folder/annotation disagreements in the current set._")
    L.append("")
    (REPO / "LABEL_AUDIT.md").write_text("\n".join(L))
    print(f"wrote LABEL_AUDIT.md ({len(mism)} mismatches)")


# ───────────────────────── misclassified_6sheets.xlsx ─────────────────────────
def misclassified_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    Z = np.load(du.ROOT / "outputs_final" / "probs_cache.npz")
    meta = json.loads((du.ROOT / "outputs_final" / "probs_cache_meta.json").read_text())
    test_paths = meta["splits"]["test"]["paths"]
    yt = np.array(meta["splits"]["test"]["y"])

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="7B1E1E")
    head_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill("solid", fgColor="F4CCCC")

    # summary sheet
    ws = wb.active; ws.title = "Summary"
    ws.append(["Model", "Test N", "Errors", "Accuracy", "False neg (missed unsafe)", "False pos"])
    for c in ws[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal="center")
    for m in ORDER:
        p = Z[f"p_{SHORT[m]}_test"]
        thr = MODELS[m]["operating_thresholds"]["threshold_balanced"]
        pred = (p >= thr).astype(int)
        err = int((pred != yt).sum())
        fn = int(((pred == 0) & (yt == 1)).sum()); fp = int(((pred == 1) & (yt == 0)).sum())
        ws.append([m, len(yt), err, round(1 - err/len(yt), 3), fn, fp])

    # per-model sheets
    def vehicle_of(p): return Path(p).parent.parent.name
    for m in ORDER:
        p = Z[f"p_{SHORT[m]}_test"]
        thr = MODELS[m]["operating_thresholds"]["threshold_balanced"]
        pred = (p >= thr).astype(int)
        title = SHORT[m][:28]
        sh = wb.create_sheet(title=title)
        sh.append(["Image", "Vehicle", "True", "Predicted", "P(unsafe)", "Error type"])
        for c in sh[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal="center")
        for i, path in enumerate(test_paths):
            if pred[i] != yt[i]:
                etype = "missed unsafe (FN)" if yt[i] == 1 else "false alarm (FP)"
                row = [Path(path).name, vehicle_of(path),
                       "unsafe" if yt[i] == 1 else "safe",
                       "unsafe" if pred[i] == 1 else "safe", round(float(p[i]), 3), etype]
                sh.append(row)
                for c in sh[sh.max_row]:
                    c.fill = bad_fill
        for col in sh.columns:
            w = max(len(str(c.value)) for c in col) + 2
            sh.column_dimensions[col[0].column_letter].width = min(w, 40)

    # make the .xlsx byte-reproducible (otherwise its embedded timestamps churn
    # the tracked binary on every rerun -> spurious git diffs)
    import datetime
    fixed = datetime.datetime(2026, 6, 30, 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    out = REPO / "misclassified_6sheets.xlsx"
    wb.save(out)
    _normalize_xlsx(out)            # fix zip member timestamps -> byte-reproducible
    print(f"wrote {out.name} ({len(wb.sheetnames)} sheets)")


def _normalize_xlsx(path):
    """Rewrite the .xlsx zip with fixed member timestamps + sorted entries so
    repeated runs produce byte-identical files (no spurious git churn)."""
    import zipfile
    fixed_dt = (1980, 1, 1, 0, 0, 0)
    with zipfile.ZipFile(path, "r") as zin:
        items = sorted(zin.infolist(), key=lambda i: i.filename)
        members = [(i.filename, zin.read(i.filename)) for i in items]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in members:
            zi = zipfile.ZipInfo(name, date_time=fixed_dt)
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = 0o600 << 16
            zout.writestr(zi, data)


if __name__ == "__main__":
    report_md(); label_audit_md()
    # NOTE: the tracked misclassified_6sheets.xlsx (6-person error-analysis,
    # val+test, humanized + thumbnails) is built by rt_group6.py — run that.
