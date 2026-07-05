"""Build the 6-person error-analysis workbook from the MISCLASSIFIED val+test set.

- Uses the already-computed val+test predictions of all 9 models
  (outputs_final/probs_cache.npz; prediction = P(unsafe) >= that model's
  tuned BALANCED threshold).
- Keeps ONLY the images at least one model got wrong (65 of the 158 val+test
  images).
- Splits those 65 across 6 people. Asif and Tanzila get the fewest rows;
  the other four share the rest. The pick order runs over a category-sorted
  list so every sheet still gets a balanced mix of val/test x vehicle x class.
- One sheet per person, IDENTICAL columns, an embedded thumbnail per row,
  wrong-model cells shaded red / correct green, and the Reasoning column filled
  with a one-line cause and a one-line fix that were written after looking at
  every image.

Writes: repo/misclassified_6sheets.xlsx
"""
from __future__ import annotations
import json
import shutil
from collections import Counter
from pathlib import Path

import numpy as np
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import du

REPO = du.ROOT.parent
OUTF = du.ROOT / "outputs_final"
THUMBS = du.ROOT / "outputs_final" / "_thumbs"
MET = json.loads((OUTF / "metrics_full.json").read_text())["models"]
Z = np.load(OUTF / "probs_cache.npz")
META = json.loads((OUTF / "probs_cache_meta.json").read_text())

MEMBERS = ["asif", "tanzila", "taif", "amio", "tazkia", "walid"]
LOW = ["asif", "tanzila"]        # these two get the fewest rows
# (column header, display model name, npz key)
MODELS = [("Ensemble", "Ensemble (best)", "ensemble"), ("ResNet50", "ResNet50", "resnet50"),
          ("ConvNeXt-Tiny", "ConvNeXt-Tiny", "convnext"), ("EfficientNet-B0", "EfficientNet-B0", "efficientnet"),
          ("ResNet18", "ResNet18", "resnet18"), ("CNN", "CNN", "cnn"),
          ("SVM", "SVM (RBF)", "svm"), ("Logistic Regression", "Logistic Regression", "logreg"),
          ("Naive Bayes", "Naive Bayes", "nb")]
VEH = {"bus": "Bus", "legua": "Leguna"}
SPLIT_NAME = {"val": "Validation", "test": "Test"}

# Per-image cause + fix, written after viewing all 65 misclassified images.
# reason = why the model(s) got it wrong ; fix = one concrete way to fix it.
REASONS = {
    # ---- Bus, true SAFE, models cried unsafe (false alarm) ----
    "IMG_3493": ("A man stands in the open rear doorway of a double-decker in heavy traffic, so almost every model read the filled door as a hanger.",
                 "Add more single-rider-in-the-doorway safe shots taken in busy traffic."),
    "IMG_3494": ("One man stands on the rear step at the open door, which looks the same as hanging even though the bus is stopped.",
                 "Add lone-rider-on-the-step safe examples."),
    "IMG_3500": ("A woman crosses in front of two buses in traffic and the linear model scored the busy scene, not a door.",
                 "Retrain logistic regression on vehicle crops so crossing pedestrians do not raise the score."),
    "IMG_3504": ("A single man fills the open rear doorway of a parked double-decker, which the models read as a hanger.",
                 "Add safe examples of one passenger standing in the doorway."),
    "IMG_3505": ("A man stands on the rear step at the open door and the deep models read the filled step as hanging.",
                 "Add more single-rider-on-the-step safe examples for the deep models."),
    "IMG_3508": ("The bus rear is empty but a cyclist passes alongside, and one backbone tripped on the busy street.",
                 "Add empty-rear safe examples with people passing."),
    "IMG_3520": ("A man stands at the open rear door with a rickshaw in front, which most models read as a hanger.",
                 "Add safe shots of a lone rider near the rear door in traffic."),
    "IMG_3523": ("A person stands on the step at the open rear door while a man walks past, which one backbone read as hanging.",
                 "Add lone-rider-on-the-step safe examples for that backbone."),
    "IMG_3525": ("A woman stands on the ground right behind the bus and the models read her as clinging to the rear.",
                 "Add more shots of people standing behind the rear."),
    "IMG_3530": ("A man stands on the rear step with his body slightly out, which the two weakest models read as hanging.",
                 "Add lone-rider-on-the-step safe examples for the classical models."),
    "IMG_3540": ("One man fills the open front doorway, which one backbone read as a person hanging out.",
                 "Add safe examples of a single passenger standing in the doorway."),
    "IMG_3541": ("A traffic jam fills the frame with the bus half hidden, so the weaker models scored the clutter.",
                 "Retrain the classical and small models on vehicle crops."),
    "IMG_3547": ("A woman stands on the step in the open doorway leaning slightly out, visually the same as a hanger, so most models flagged it.",
                 "Add near-identical safe and unsafe doorway pairs so the fine cue can be learned."),
    "IMG_3582": ("A lone man stands at the open rear door, which only the color model read as hanging.",
                 "Add more single-rider-at-the-door safe examples for Naive Bayes."),
    "IMG_3596": ("A man stands beside the bus at the open side door with passengers seated inside, which reads as door activity.",
                 "Add safe images of an open door with a seated load and a bystander alongside."),
    "IMG_3599": ("The rear door is open and a woman walks behind the bus, which the models read as boarding even though the bus is empty.",
                 "Add safe examples of an open rear door with people passing behind."),
    "IMG_3620": ("A rickshaw van crosses in front and a man stands in a second bus doorway, which one backbone read as hanging.",
                 "Add busy-street safe examples with a rider in a background doorway."),
    "IMG_3655": ("A man leans out of the open front doorway at a stop, which the linear model read as hanging.",
                 "Add safe examples of a passenger leaning in the doorway for logistic regression."),
    "IMG_3754": ("A person stands at the rear door with a woman in the foreground, which the weaker models read as door activity.",
                 "Add safe rear-door examples with a foreground pedestrian."),
    "IMG_3763": ("The rear is empty but rickshaws crowd the left, and two models scored the clutter, not the bus.",
                 "Add empty-rear safe examples in dense traffic."),
    "IMG_3779": ("A man stands in the open front doorway with a roadside crowd behind, which one backbone read as hanging.",
                 "Add lone-rider-in-doorway safe examples for that backbone."),
    "IMG_3955": ("A man stands in the open doorway and a backpacker passes in front, which two weaker models read as door activity.",
                 "Add safe doorway examples with a foreground pedestrian."),
    "IMG_3971": ("A rider stands on the doorway step with his body toward the opening, almost identical to a hanger, so the strong models flagged it.",
                 "Add near-identical safe and unsafe doorway pairs."),
    # ---- Bus, true UNSAFE, models missed it ----
    "IMG_3534": ("The hanger sits in the open side door behind heavy foreground traffic, so the SVM lost him in the clutter.",
                 "Retrain the SVM on vehicle crops with more low-visibility hangers."),
    "IMG_3535": ("A woman leans out of the door among a boarding crowd at a stop, which the classical models read as ordinary boarding.",
                 "Add crowded-stop hanging examples for the classical models."),
    "IMG_3577": ("A woman leans well out of the front doorway of an almost empty bus, but the small figure on a clean bus was under-weighted by most models.",
                 "Add single-hanger examples on otherwise empty buses."),
    "IMG_3578": ("A woman stands on the step leaning out with a child, but she is small against a clean bus, so several models missed her.",
                 "Add more small-hanger-on-a-clean-bus examples."),
    "IMG_3615": ("A man hangs at the rear opening while students board with backpacks, and the weaker backbones lost the small figure in the crowd.",
                 "Add crowded-stop hanging examples for those backbones."),
    "IMG_3778": ("The door figure is small and boxed in by a jam, so the lighter models read a plain bus.",
                 "Crop to the vehicle and add more hangers photographed in jams."),
    "IMG_4017": ("The hanger is a small figure lost in a dense jam of buses and rickshaws, so the linear model saw clutter.",
                 "Retrain logistic regression on vehicle crops."),
    "IMG_4018": ("A rider clearly hangs on the rear footboard, but heavy foreground motorbike traffic dominates the frame and every model keyed on that instead.",
                 "Add examples where the vehicle sits behind foreground traffic so the model attends to the bus."),
    "IMG_4029": ("A man with a backpack stands on the rear step among a boarding crowd, which the models read as routine boarding.",
                 "Add terminal boarding-against-hanging pairs."),
    "IMG_4109": ("A man and a woman ride the footboard at the open door, missed only by the linear model.",
                 "Add more footboard-hanging examples for logistic regression."),
    "IMG_4110": ("The same footboard riders as the previous frame, again missed only by the linear model.",
                 "Add more footboard-hanging examples for logistic regression."),
    "IMG_4134": ("A man leans out of the open door while women cross in front, and the two weakest models scored the foreground crossing.",
                 "Add hanging examples with foreground pedestrians for the classical models."),
    "IMG_4176": ("A man rides at the open rear door boxed in by cars, and the two weak models read him as boarding.",
                 "Add rear-door hanging examples in dense traffic for the small and color models."),
    "IMG_4180": ("A man rides the footboard at the open door between two buses, which the classical models read as boarding.",
                 "Add footboard-hanging examples for the classical models."),
    "IMG_4181": ("The same footboard rider as the previous frame, again missed by the three classical models.",
                 "Add more footboard-hanging examples for the classical models."),
    "IMG_4189": ("Riders press at the open door inside a dense boarding crowd, which the two weakest models read as a normal queue.",
                 "Add crowded-stop hanging examples for the classical models."),
    "IMG_4194": ("A man leans out of the door above a boarding crowd, which the classical models lost in the crowd.",
                 "Add crowded-stop hanging examples for the classical models."),
    # ---- Leguna, true SAFE, models cried unsafe ----
    "IMG_3682": ("A man stands on the rear step of a leguna full of seated passengers, which almost every model read as hanging.",
                 "Add safe examples of a passenger boarding the rear step at a stop."),
    "IMG_3699": ("A seated passenger reaches up to the roof rail, which the linear model read as hanging.",
                 "Add seated-with-arm-on-rail safe examples for logistic regression."),
    "IMG_3792": ("The open back is full of seated women with legs near the edge, which one backbone read as hanging.",
                 "Add full-but-seated leguna examples."),
    "IMG_3808": ("A man bends into the open rear to enter, which the SVM read as riding the step.",
                 "Add safe boarding-at-the-rear examples for the SVM."),
    "IMG_3810": ("Passengers sit at the open edge of the leguna, which the two weak models read as edge riders.",
                 "Add safe examples of passengers seated at the open edge."),
    "IMG_3811": ("A passenger sits at the open edge of the leguna, which the classical models read as hanging.",
                 "Add edge-seated safe examples for the classical models."),
    "IMG_3828": ("A passenger sits at the rear edge with his legs out, which the color model read as hanging.",
                 "Add edge-seated safe examples for Naive Bayes."),
    "IMG_3830": ("The open back is empty but veiled women stand in the foreground, and the classical models scored them.",
                 "Add empty-open-back safe examples with foreground pedestrians."),
    "IMG_3842": ("Women sit near the open edge with a man in the foreground, which the linear model read as door activity.",
                 "Add edge-seated safe examples for logistic regression."),
    "IMG_3866": ("A night shot of a leguna with a passenger's feet near the open rear, which reads as hanging in the dark.",
                 "Add low-light crowded-but-seated examples."),
    "IMG_3872": ("The open back is empty but a man stands in the foreground, and the two weak models scored him.",
                 "Add empty-open-back safe examples with a bystander alongside."),
    "IMG_3958": ("The leguna carries seated passengers with men crowding the foreground, which the models read as edge riders.",
                 "Add full-but-seated leguna examples with foreground pedestrians."),
    "IMG_3961": ("An empty open back parked at a railing, where the two weak models flagged the open shape alone.",
                 "Add empty-open-back safe examples so the shape alone is not flagged."),
    "IMG_3965": ("A man stands on the ground beside the leguna with a passenger seated at the edge, which the two weak models read as hanging.",
                 "Add safe examples with a person standing beside the vehicle."),
    # ---- Leguna, true UNSAFE, models missed it ----
    "IMG_3564": ("A man rides the rear step with his back to the camera, which the color model read as boarding.",
                 "Add rear-step hanging examples for Naive Bayes."),
    "IMG_3573": ("A man leans in at the open side holding the rail, which the classical models read as boarding rather than riding.",
                 "Add side-riding examples for the classical models."),
    "IMG_3707": ("A man stands on the rear step on his phone in a calm pose, which the small CNN read as waiting rather than riding.",
                 "Add rear-step hanging examples with relaxed poses for the CNN."),
    "IMG_3878": ("A woman poses holding the frame with one foot on the step, which the two weak models read as boarding.",
                 "Add footboard-riding examples with a boarding-like pose."),
    "IMG_3879": ("The same posed rider as the previous frame, again missed by the color model.",
                 "Add more footboard-riding examples with a boarding-like pose for Naive Bayes."),
    "IMG_3933": ("A man rides the rear step with his back to the camera, which the color model read as boarding.",
                 "Add rear-step hanging examples for Naive Bayes."),
    "IMG_3936": ("Two men ride the rear step with backpacks, which the two classical models read as boarding passengers.",
                 "Add two-person rear-step hanging examples for the classical models."),
    "IMG_4206": ("Riders stand on the rear step at night, which the two weak models lost in the low light.",
                 "Add low-light rear-step hanging examples for the classical models."),
    "IMG_4216": ("A man climbs into the rear at dusk carrying a bag, which the color model read as loading cargo.",
                 "Add low-light footboard-riding examples for Naive Bayes."),
    "IMG_4220": ("A man rides the rear step at night among a crowd, which the two weak models missed in the low light.",
                 "Add low-light rear-step hanging examples for the classical models."),
    "IMG_4221": ("A man climbs onto the rear step at night, which the color model read as boarding in the dark.",
                 "Add low-light rear-step hanging examples for Naive Bayes."),
}


def quotas(n):
    """Row counts per person; Asif and Tanzila get the fewest, rest shared."""
    low = n // 8
    rest = n - 2 * low
    base = rest // 4
    extra = rest - base * 4
    q = {"asif": low, "tanzila": low, "taif": base, "amio": base, "tazkia": base, "walid": base}
    for i in range(extra):                        # spread any remainder over the big four
        q[["taif", "amio", "tazkia", "walid"][i]] += 1
    return q


def assign(recs, q):
    """Greedy proportional pick over a category-sorted list so each sheet keeps a
    balanced val/test x vehicle x class mix while honouring the per-person quota."""
    remaining = dict(q)
    groups = {m: [] for m in MEMBERS}
    for r in recs:
        pick = max((m for m in MEMBERS if remaining[m] > 0),
                   key=lambda m: (remaining[m] / q[m], remaining[m]))
        groups[pick].append(r); remaining[pick] -= 1
    return groups


def transfer(groups, donor, recipient):
    """Move one image donor -> recipient, preferring an unsafe (then leguna) one so
    the safe/bus-heavy recipient sheets even out."""
    r = max(groups[donor], key=lambda r: (r["true"] == "Unsafe", r["vehicle"] == "Leguna", -r["imgnum"]))
    groups[donor].remove(r); groups[recipient].append(r)


def main():
    THUMBS.mkdir(parents=True, exist_ok=True)
    thr = {h: MET[full]["operating_thresholds"]["threshold_balanced"] for h, full, _ in MODELS}

    # every val+test image the models disagree with the label on (>=1 model wrong)
    recs = []
    for sp in ("val", "test"):
        paths = META["splits"][sp]["paths"]; y = META["splits"][sp]["y"]
        for i, p in enumerate(paths):
            pth = Path(p); stem = pth.stem
            true = "Unsafe" if y[i] == 1 else "Safe"
            preds, wrong = {}, []
            for h, full, key in MODELS:
                pred = "Unsafe" if float(Z[f"p_{key}_{sp}"][i]) >= thr[h] else "Safe"
                preds[h] = pred
                if pred != true:
                    wrong.append(h)
            if not wrong:
                continue                                   # keep only misclassified
            reason, fix = REASONS.get(stem, ("", ""))
            recs.append({"stem": stem, "vehicle": VEH.get(pth.parent.parent.name, pth.parent.parent.name),
                         "split": SPLIT_NAME[sp], "true": true, "preds": preds,
                         "wrong": wrong, "path": p, "reason": reason, "fix": fix,
                         "imgnum": int("".join(ch for ch in stem if ch.isdigit()) or 0)})

    missing = [r["stem"] for r in recs if not r["reason"]]
    if missing:
        raise SystemExit(f"no reasoning written for: {missing}")

    # category-sorted, then quota-aware balanced assignment (Asif/Tanzila lowest)
    recs.sort(key=lambda r: (r["split"], r["vehicle"], r["true"], r["imgnum"]))
    q = quotas(len(recs))
    groups = assign(recs, q)
    # bump Asif & Tanzila to 10 each: pull two into Asif (from Taif, Amio) and two
    # into Tanzila (from Tazkia, Walid). Others settle at 11 each.
    for donor, recipient in [("taif", "asif"), ("amio", "asif"), ("tazkia", "tanzila"), ("walid", "tanzila")]:
        transfer(groups, donor, recipient)

    # styles
    hdr_fill = PatternFill("solid", fgColor="7B1E1E"); hdr_font = Font(bold=True, color="FFFFFF", size=11)
    wrong_fill = PatternFill("solid", fgColor="F6C6C0"); right_fill = PatternFill("solid", fgColor="CFE8CF")
    safe_font = Font(color="1B5E20", bold=True); unsafe_font = Font(color="8C1D0B", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E0CDB4"); border = Border(thin, thin, thin, thin)

    cols = ["Photo", "Image", "Vehicle", "Dataset", "True Label"] + [h for h, _, _ in MODELS] + \
           ["Models That Got It Wrong", "# Wrong", "Reasoning / Notes"]

    wb = Workbook(); wb.remove(wb.active)
    for name in MEMBERS:
        ws = wb.create_sheet(name.capitalize())
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c); cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border
        ws.row_dimensions[1].height = 30
        for r in groups[name]:
            note = f"{r['reason']}\nFix: {r['fix']}"
            row = ["", f"{r['stem']}.png", r["vehicle"], r["split"], r["true"]] + \
                  [r["preds"][h] for h, _, _ in MODELS] + \
                  [", ".join(r["wrong"]), len(r["wrong"]), note]
            ws.append(row)
            ridx = ws.max_row
            ws.row_dimensions[ridx].height = 96
            # embed thumbnail
            try:
                tp = THUMBS / f"{r['split']}_{r['vehicle']}_{r['stem']}.png"
                if not tp.exists():
                    im = PILImage.open(r["path"]).convert("RGB"); im.thumbnail((130, 130))
                    im.save(tp)
                xim = XLImage(str(tp)); xim.width, xim.height = 92, 92
                ws.add_image(xim, f"A{ridx}")
            except Exception:
                ws.cell(row=ridx, column=1, value="(no image)")
            # formatting + coloring
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=ridx, column=c); cell.border = border
                cell.alignment = left if c == len(cols) else center
            tl = ws.cell(row=ridx, column=5); tl.font = unsafe_font if r["true"] == "Unsafe" else safe_font
            for j, (h, _, _) in enumerate(MODELS):
                cell = ws.cell(row=ridx, column=6 + j)
                pred = r["preds"][h]
                cell.fill = wrong_fill if pred != r["true"] else right_fill
                cell.font = unsafe_font if pred == "Unsafe" else safe_font
        # widths + freeze header + autofilter
        widths = [15, 15, 10, 12, 11] + [13] * len(MODELS) + [26, 9, 52]
        for c, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = wdt
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = f"B1:{get_column_letter(len(cols))}1"

    out = REPO / "misclassified_6sheets.xlsx"
    wb.save(out)

    # one image folder per person (full-resolution copies, file name = xlsx Image column)
    root = REPO / "misclassified_by_person"
    if root.exists():
        shutil.rmtree(root)
    for name in MEMBERS:
        d = root / name.capitalize(); d.mkdir(parents=True, exist_ok=True)
        for r in groups[name]:
            shutil.copy2(r["path"], d / f"{r['stem']}.png")

    print(f"wrote {out.name}  ({len(wb.sheetnames)} person-sheets, identical columns)")
    print(f"wrote {root.name}/ (one folder per person with their images)")
    print(f"pool = {len(recs)} misclassified val+test images (>=1 model wrong)  ·  reasoning filled")
    print("\nper-person assignment (Asif & Tanzila lowest):")
    for name in MEMBERS:
        g = groups[name]
        c = Counter(r["true"] for r in g); s = Counter(r["split"] for r in g)
        v = Counter(r["vehicle"] for r in g)
        print(f"  {name:8s}: {len(g):2d} imgs | Safe {c['Safe']:2d} Unsafe {c['Unsafe']:2d} | "
              f"Val {s['Validation']:2d} Test {s['Test']:2d} | Bus {v['Bus']:2d} Leguna {v['Leguna']:2d}")


if __name__ == "__main__":
    main()
