"""Fill the Reasoning column for amio and taif sheets with vision-based, humanized
explanations (why each image was misclassified + how to fix), keyed by row."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

XLSX = Path(__file__).resolve().parent.parent.parent / "error_analysis" / "misclassified_6sheets.xlsx"

AMIO = {
 2: "Night photo of a leguna from behind — passengers are SEATED inside, nobody is hanging, so 'safe' is correct. Only Logistic Regression failed: on a dark, grainy 128-px grey HOG image the bright shop signs and tangled overhead wires form strong edges, and HOG cannot tell seated riders from a hanger. Fix: add night/low-light brightness augmentation and CLAHE contrast-normalisation before HOG; collect a few more night frames.",
 3: "A pink bus parked in daylight with an empty doorway — clearly safe. Logistic Regression alone trips because on grey-scale HOG the dark door opening plus tree/shadow edges resemble the gradient pattern of a person standing in a doorway. Fix: give the classical models more 'open door, nobody hanging' bus shots and add colour features so shadows aren't read as bodies.",
 4: "Another night leguna frame (burst shot), passengers seated inside — safe. Logistic Regression is again fooled by the low light and bright signage/wires behind the vehicle. Fix: same night augmentation + contrast-normalisation; the deep models already handle these, so it is purely a weak-feature/low-light problem.",
 5: "Pink bus parked side-on, door open but no one hanging out — safe. The edge-based HOG behind Logistic Regression reads the open door and its shadow as a figure. Fix: more side-view 'door open / nobody hanging' samples; this is a feature limitation, not a data shortage, since every deep model gets it right.",
 6: "Colourful bus in daylight; a conductor stands just INSIDE the open door with feet on the step — not hanging out, so labelled safe. EfficientNet-B0 over-reacts to a person being at the doorway. This is a genuine 'standing in the doorway (safe) vs leaning/hanging outside (unsafe)' borderline. Fix: label matched pairs of these two cases so the boundary is learned — the most valuable examples to add.",
 7: "A passenger is on the entry step / hanging at the bus door — unsafe. Only the from-scratch CNN misses it; at 128 px the small figure at the door is washed out for a shallow network while the ImageNet-pretrained models still catch it. Fix: train the CNN at higher resolution (224) and/or add door-region crops, plus more close-range hanging examples.",
 8: "Pink bus shot from the rear, parked, no passengers — safe. Logistic Regression is confused by the busy rear (tail-lights, Bangla text, peeling paint) whose edges mimic a human shape. Fix: add more rear-view empty-bus images and colour/context features for the classical pipeline.",
 9: "A passenger hangs at the FAR door of the blue bus, but the figure is small and partly hidden by the bus body and street clutter — so seven of eight models miss it (only Logistic Regression calls it unsafe, essentially by chance). A hard 'small/distant hanger' case. Fix: higher-resolution input and door-region crop training, more distant-hanger photos; also worth a second human label check.",
 10: "A red double-decker far across a crowded intersection, labelled unsafe — but the hanging passenger is essentially too small/occluded to see at training resolution, so EVERY model gets it wrong. The hardest image in the set: either an extreme distant case or a borderline label. Fix: flag for label review; if truly unsafe it needs zoomed/high-resolution detection rather than whole-image classification.",
 11: "Green bus in mid-distance with rickshaws crowding the foreground; a passenger hangs at the door — unsafe. Only the shallow CNN misses it: the busy foreground plus small subject overwhelm its low capacity, while the transfer models succeed. Fix: higher resolution and foreground-robust augmentation (random occluders), more cluttered-scene examples.",
 12: "An empty leguna from behind — no passengers, nobody hanging, safe. Logistic Regression trips on the bright yellow/red frame and 'STOP' lettering whose strong edges look 'busy' to HOG. Fix: add more empty-vehicle samples and apply contrast normalisation for the HOG models.",
 13: "A leguna with passengers SEATED inside the open cabin — safe. SVM reads the visible seated figures at the opening as 'people at the door' and calls it unsafe — the classic inside-vs-hanging confusion for a HOG model. Fix: add many more 'seated passengers visible through the opening = safe' examples.",
 14: "Same pattern — a leguna full of seated passengers with open sides showing faces; safe. Both classical models false-alarm on the visible people because HOG cannot tell seated-inside from hanging-outside. Fix: this 'open-sided leguna with seated riders' pattern is under-represented; add more of it (the deep models already cope).",
 15: "A small leguna/pickup from behind with a passenger seated inside the rear (legs visible) but fully INSIDE — safe. Logistic Regression sees the partial figure at the open back and flips. Fix: add rear-open-vehicle examples; colour/depth cues would help the classical model separate 'inside' from 'hanging'.",
}

TAIF = {
 2: "A dark-coloured bus in daylight, nobody hanging — safe. Logistic Regression is thrown by the dark bus body together with a pedestrian in the foreground, whose combined edges resemble a person at the vehicle. Fix: more clean bus examples and colour features; the deep models already classify this correctly.",
 3: "Green/orange bus approaching with a cyclist and pedestrians in front — nobody hanging, safe. The cluttered foreground confuses HOG/Logistic Regression, which blends the bus-front pattern with the nearby people. Fix: hard-negative training with busy foregrounds; colour/context features.",
 4: "A passenger hangs at the door of the green/yellow bus in mid-distance — unsafe. Only the from-scratch CNN misses it because the small, distant figure is lost at 128 px. Fix: higher-resolution CNN input and door-region crops, plus more distant-hanger examples.",
 5: "An empty leguna from behind at dusk — nobody inside or hanging, safe. Both classical models false-trigger: the dark interior plus bright yellow frame edges in low light look like a figure to HOG. Fix: night/low-light augmentation, more empty-vehicle samples, and contrast-normalise before HOG.",
 6: "Night scene; a leguna whose doorway has a person boarding/seated (not hanging out) with a bus alongside — labelled safe. Logistic Regression calls it unsafe because in the dark a person near the opening reads as a hanger. Fix: night data plus clear 'boarding/seated = safe' vs 'hanging = unsafe' examples.",
 7: "A man clearly stands on the rear step of the yellow leguna, body outside and holding on — unambiguously unsafe and well-lit. Surprisingly only the from-scratch CNN misses it (predicts safe), showing the shallow 128-px network hasn't learned this posture robustly. Fix: add exactly these clear examples for the CNN and train it at higher resolution.",
 8: "Burst frame of the same man hanging on the back of the yellow leguna — clearly unsafe. The CNN again predicts safe, confirming a real weakness of the from-scratch model on the hanging posture (not an image-quality issue). Fix: more such examples + higher-resolution CNN, or replace it with a pretrained backbone.",
 9: "Night shot of a leguna with a person SEATED in the doorway (not hanging) and a bus passing — safe. SVM false-alarms on the figure visible at the opening in low light. Fix: night augmentation and more 'seated in doorway = safe' examples.",
 10: "Daytime — a man with a backpack standing on the rear step of the leguna, body outside — clearly unsafe. (Same image NUMBER as the night 'safe' frame above, but a different photo.) Only the CNN misses it. Fix: higher-resolution CNN and more hanging examples; also de-duplicate/rename this clashing file name in the dataset.",
 11: "Another clear daytime hanger — backpacked man standing on the leguna's back step, unsafe. The from-scratch CNN predicts safe while all transfer models get it right. Fix: the CNN needs higher resolution and more of these clear hanging postures.",
 12: "Close rear view of an EMPTY, rusty leguna interior — no people, safe. Logistic Regression is fooled because peeling paint, seats and the metal frame create dense edges/texture that HOG reads as a figure. Fix: more empty-vehicle examples and contrast normalisation; consider a texture-suppressing pre-filter for HOG.",
 13: "LABEL ERROR: this image is an EMPTY parked leguna seen from behind — there is NO passenger hanging, so the true situation is SAFE, but the dataset labels it 'unsafe'. The CNN's 'safe' prediction is actually correct; the label is wrong. Fix: correct this label to safe (or remove the image). This mislabelled case is also what blocks 100% recall elsewhere — see LABEL_AUDIT.md.",
 14: "An empty leguna parked side-on, open side showing seats — nobody hanging, safe. Three strong models (ResNet18/ConvNeXt/EffNet) false-positive, probably mistaking the vertical door pillar/frame for a standing person. Fix: add this as a hard-negative ('empty open-sided leguna = safe'); a handful of such images should fix the deep models.",
 15: "Empty leguna from behind — safe. ResNet18 alone false-triggers, again likely mistaking the interior frame/pole for a person. Fix: more empty-leguna hard negatives; a rare slip since the other deep models get it right.",
}

wb = load_workbook(XLSX)
RCOL = None
ws0 = wb["amio"]
for c in range(1, ws0.max_column + 1):
    if ws0.cell(1, c).value == "Reasoning":
        RCOL = c
assert RCOL, "Reasoning column not found"

for sheet, texts in (("amio", AMIO), ("taif", TAIF)):
    ws = wb[sheet]
    for r, txt in texts.items():
        cell = ws.cell(r, RCOL)
        cell.value = txt
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[ws.cell(1, RCOL).column_letter].width = 90
    for r in texts:
        ws.row_dimensions[r].height = 95
    print(f"{sheet}: filled {len(texts)} reasoning cells")

wb.save(XLSX)
print(f"saved {XLSX}")
