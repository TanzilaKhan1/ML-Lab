"""Take the images misclassified by >=1 model (excluding Naive Bayes), sort them,
split across 6 people, RE-RUN the 8 models (no NB), and write an Excel workbook
with 6 sheets (one per person).

Sheet columns: Image (IMG_xxxx.png) | Type (bus/leguna) | True label |
               <8 model predictions safe/unsafe> | Models misclassified | Reasoning
Cells where a model is WRONG are shaded red; correct = green.
"""
from __future__ import annotations

import csv
import re
import shutil
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
PROJECT = REPO.parent
sys.path.insert(0, str(REPO / "predictor"))
from predictor_app import inference as inf  # noqa: E402
from predictor_app.preprocess import preprocess_for_model, standardize_image  # noqa: E402

EA = PROJECT / "error_analysis"
FLAT = PROJECT / "data" / "dataset_safeunsafe"
GROUP = EA / "group6"
MEMBERS = ["asif", "tanzila", "taif", "amio", "tazkia", "walid"]
# 8 models (Naive Bayes excluded), SHORT -> full display name
MODELS = [("LogReg", "Logistic Regression"), ("SVM", "SVM (RBF)"), ("CNN", "CNN"),
          ("ResNet18", "ResNet18"), ("ResNet50", "ResNet50"), ("ConvNeXt", "ConvNeXt-Tiny"),
          ("EffNetB0", "EfficientNet-B0"), ("Ensemble", "Ensemble (best)")]
SHORTS = [s for s, _ in MODELS]
VEH = {"bus": "bus", "legua": "leguna"}


def p_unsafe(model, kind, feats, pil):
    if kind == "sklearn":
        if hasattr(model, "predict_proba"):
            return float(model.predict_proba(feats)[0, 1])
        s = float(np.atleast_1d(model.decision_function(feats))[0])
        return 1.0 / (1.0 + np.exp(-s))
    return float(model.predict_proba(pil)[0, 1])


def imgnum(stem):
    m = re.search(r"(\d+)", stem)
    return int(m.group(1)) if m else 0


def main():
    rows = list(csv.DictReader(open(EA / "all385_all_models_inference.csv")))
    mis = [r for r in rows if any(r[f"{s}_pred"] != r["true_label"] for s in SHORTS)]
    # sort by image number, then true_label (stable for duplicate stems)
    mis.sort(key=lambda r: (imgnum(r["image"]), r["true_label"]))
    print(f"misclassified by >=1 of 8 models (no NB): {len(mis)}")

    # partition into 6 near-equal contiguous groups
    n = len(mis); base, rem = divmod(n, len(MEMBERS))
    groups, idx = {}, 0
    for i, name in enumerate(MEMBERS):
        cnt = base + (1 if i < rem else 0)
        groups[name] = mis[idx:idx + cnt]; idx += cnt

    # load 8 models once
    loaded = {full: (inf.load_model(full), inf.AVAILABLE_MODELS[full][1]) for _, full in MODELS}

    if GROUP.exists():
        shutil.rmtree(GROUP)
    GROUP.mkdir(parents=True)

    wb = Workbook(); wb.remove(wb.active)
    hdr_fill = PatternFill("solid", fgColor="1F2937"); hdr_font = Font(bold=True, color="FFFFFF")
    wrong_fill = PatternFill("solid", fgColor="F8C9C9")   # red-ish
    right_fill = PatternFill("solid", fgColor="CDEAD4")    # green-ish
    safe_font = Font(color="166534"); unsafe_font = Font(color="991B1B")

    for name in MEMBERS:
        ws = wb.create_sheet(name)
        mdir = GROUP / name; mdir.mkdir(parents=True)
        cols = ["Image", "Type", "True label"] + SHORTS + ["Models misclassified", "Reasoning"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c); cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        used = set()
        for r in groups[name]:
            stem = r["image"].replace(".png", "").split("__", 1)[-1]   # IMG_xxxx
            true = r["true_label"]; veh = VEH.get(r["vehicle"], r["vehicle"])
            # re-run inference
            src = FLAT / true / r["image"]
            pil = standardize_image(str(src)); feats = preprocess_for_model(pil)
            preds, wrong_models = [], []
            for short, full in MODELS:
                model, kind = loaded[full]
                pu = p_unsafe(model, kind, feats, pil)
                pred = "unsafe" if pu >= 0.5 else "safe"
                preds.append(pred)
                if pred != true:
                    wrong_models.append(short)
            # copy image to folder (unique name)
            fn = f"{stem}.png"
            if fn in used:
                fn = f"{stem}_{true}.png"
            used.add(fn)
            if src.exists():
                shutil.copy2(src, mdir / fn)
            # write row
            rowvals = [f"{stem}.png", veh, true] + preds + ["+".join(wrong_models), ""]
            ws.append(rowvals)
            ridx = ws.max_row
            # color model cells
            for j, (short, pred) in enumerate(zip(SHORTS, preds)):
                cell = ws.cell(row=ridx, column=4 + j)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = wrong_fill if pred != true else right_fill
                cell.font = unsafe_font if pred == "unsafe" else safe_font
            ws.cell(row=ridx, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=ridx, column=2).alignment = Alignment(horizontal="center")
        # widths + freeze
        widths = [16, 9, 11] + [10] * len(SHORTS) + [26, 50]
        for c, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = wdt
        ws.freeze_panes = "A2"

    out = EA / "misclassified_6sheets.xlsx"
    wb.save(out)
    print("\n=== per-person split ===")
    for name in MEMBERS:
        g = groups[name]
        from collections import Counter
        c = Counter(r["true_label"] for r in g); v = Counter(VEH.get(r["vehicle"]) for r in g)
        print(f"  {name:8s}: {len(g)} imgs | safe {c['safe']} unsafe {c['unsafe']} | "
              f"bus {v['bus']} leguna {v['leguna']} | folder {GROUP/name}")
    print(f"\nExcel -> {out} (6 sheets)")
    print(f"images -> {GROUP}/<name>/")


if __name__ == "__main__":
    main()
